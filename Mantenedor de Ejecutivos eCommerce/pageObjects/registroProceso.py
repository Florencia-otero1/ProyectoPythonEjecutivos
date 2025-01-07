import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
import os

class registroProceso:

    def __init__(self, carpeta_salida, nombre_archivo):  

        self.archivo_salida = os.path.join(carpeta_salida,nombre_archivo)
        self.escribir_encabezados()

    def escribir_encabezados(self):                 
        try:
            # Intenta cargar el archivo si ya existe
            workbook = openpyxl.load_workbook(self.archivo_salida)
        except FileNotFoundError:
            # Si el archivo no existe, lo crea
            workbook = Workbook()

        # Definimos los nombres de las hojas y encabezados
        hojas = ["Resultados Altas", "Resultados Bajas", "Resultados Cambios de Roles"]
        encabezados = ["Correo", "RUT", "Estado de Ejecución", "Error"]
        encabezados_cambios_de_roles = ["Correo", "RUT", "Estado de Ejecución", "Tipo de modificación", "Error"]

        # Crear hojas y agregar encabezados si no existen ya
        for nombre_hoja in hojas:
            if nombre_hoja not in workbook.sheetnames:
                hoja = workbook.create_sheet(title=nombre_hoja)
                if nombre_hoja == "Resultados Altas" or nombre_hoja == "Resultados Bajas":                    
                    hoja.append(encabezados)  # Agrega los encabezados 
                else: 
                    hoja.append(encabezados_cambios_de_roles) #agrega los encabezados de cambios de roles


        # Borra la hoja predeterminada si está vacía y se ha creado al inicio
        if "Sheet" in workbook.sheetnames and not any(workbook["Sheet"].values):
            del workbook["Sheet"]

        workbook.save(self.archivo_salida)
        
    def escribir_resultado(self, tipo_operacion, correo, rut, estado_ejecucion, error):
        # Determina la hoja correspondiente según el tipo de operación
        if tipo_operacion.lower() == "alta":
            nombre_hoja = "Resultados Altas"
        elif tipo_operacion.lower() == "baja":
            nombre_hoja = "Resultados Bajas"
        elif tipo_operacion.lower() == "cambio de rol":
            nombre_hoja = "Resultados Cambios de Roles"
        else:
            print(f"Operación desconocida: {tipo_operacion}")
            return
        # Cargar el archivo y seleccionar la hoja correspondiente
        workbook = load_workbook(self.archivo_salida)
        hoja = workbook[nombre_hoja]

        # Agregar una nueva fila con los datos
        nueva_fila = [correo, rut, estado_ejecucion, error]
        hoja.append(nueva_fila)

        # Guardar los cambios
        workbook.save(self.archivo_salida)

    def escribir_resultado_modificacion(self, tipo_operacion, correo, rut, estado_ejecucion, tipo_modificacion, error):
        # Determina la hoja correspondiente según el tipo de operación
        if tipo_operacion.lower() == "alta":
            nombre_hoja = "Resultados Altas"
        elif tipo_operacion.lower() == "baja":
            nombre_hoja = "Resultados Bajas"
        elif tipo_operacion.lower() == "cambio de rol":
            nombre_hoja = "Resultados Cambios de Roles"
        else:
            print(f"Operación desconocida: {tipo_operacion}")
            return

        # Cargar el archivo y seleccionar la hoja correspondiente
        workbook = load_workbook(self.archivo_salida)
        hoja = workbook[nombre_hoja]

        # Agregar una nueva fila con los datos
        nueva_fila = [correo, rut, estado_ejecucion, tipo_modificacion, error]
        hoja.append(nueva_fila)

        # Guardar los cambios
        workbook.save(self.archivo_salida)


    
        